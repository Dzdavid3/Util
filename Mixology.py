from openpyxl import load_workbook
import xlsxwriter
import time
import datetime
import string
d = dict(enumerate(string.ascii_uppercase, 1))


class Mixology:
    def __init__(self):
        d = datetime.datetime.now()
        self.daterann = "%i_%i_%i" % (d.year, d.month, d.day)
        self.accelerant_sources = {}  # {Source:[symbol,symbol], Source2:[symbols3, symbol4, symbol5]}
        self.path = "C:\\auto_ana\\file_2017_10_17_roi.txt"
        self.entries = {}
        self.date_ran = None
        self.starting_dollar_amount = 0.0

        # For IB Data Set
        self.ib_entries = {}
        self.starting_dollar_amount = None
        self.ib_total_money_entry = 0.0
        self.ib_total_money_exit = 0.0
        self.ib_overall_roi = 0.0

        # Rows of Data
        self.rows = []


    def get_data_from_accelerant_tracker(self):
        wb = load_workbook(filename = "C:\\Users\\Niska\\Desktop\\Mixology\\AT_10-17-17.xlsx")
        ws = wb.active

        #print(sheet_ranges['C2'].value)
        #print(wb.sheetnames)

        for column_id, col in enumerate(ws.iter_cols(min_row=2, min_col=2, max_col=16, max_row=100)):
            real_column_letter =d[(column_id +2)]
            column_number = column_id + 1
            wereat = str(real_column_letter) + "1"
            column_title = ws[wereat].value

            if column_title not in self.accelerant_sources:
                self.accelerant_sources[column_title] = []

            for cell in col:
                if cell.value is not None:
                    self.accelerant_sources[column_title].append(str(cell.value))

    def time_to_secs(self, msg):
        hours = int(msg[0] + msg[1])
        mins = int(msg[2] + msg[3])
        secs = int(msg[4] + msg[5])
        mili = float("0." + (msg.split("."))[1])
        return (hours * 3600) + (mins * 60) + secs + mili - 34200

    def process_msg_entry(self, msg):
        msg_split = msg.split(",")
        index = len(msg_split) - 1
        for num, i in enumerate(msg_split):
            if "human_time_at_notification" in i:
                human_entry_time = i[39:47]

            if "reasons_for_entry" in i:
                reason_for_entry = i[19:] + msg_split[num + 1] + msg_split[num + 2] + msg_split[num + 3] + msg_split[num + 4] + msg_split[num + 5]
                if "ts=" in msg_split[num + 6]:
                    pass
                else:
                    reason_for_entry += msg_split[num + 6]


        syl = (msg_split[index])[8:]
        symbol = filter(str.isalpha, syl)
        tim = (msg_split[index - 1])[23:]
        price = (msg_split[index - 3])[28:]

        tim = float(tim) - 34200

        #tim = float(tim) - 34200

        price = float(price)

        if symbol in self.entries:
            # ToDO: Handel Multiple Entries
            entity = str(symbol) + str(1)
            self.entries[symbol] = {}
            self.entries[symbol].update({'Entry Time': tim, "Entry Price": price})
        else:
            self.entries[symbol] = {}
            self.entries[symbol].update({'Entry Time': tim, "Entry Price": price, "Human Entry Time": human_entry_time, "Reason for Entry":reason_for_entry})

    def process_msg_exit(self, msg):
        if "EventSignalExitPosition" in msg:
            msg_split = msg.split(",")
            index = len(msg_split) - 1

            for i, msg in enumerate(msg_split):
                if "scan_info" in msg:
                    msg_split2 = msg.split("-")
                    scan_tim = msg_split2[2][3:19]
                    timy = scan_tim.replace(":", "")
                    human_scan_time = msg_split2[2][3:11]
                    scan_time = self.time_to_secs(timy)
                elif "reason_for_exit" in msg:
                    reason = msg[19:]
                elif "human_time" in msg:
                    human_exit_time = msg[23:31]
                elif "high_since_entry" in msg:
                    exclude_last = len(msg) - 1
                    highest_price_since_entry = msg[20:exclude_last]

            symbol = (msg_split[6])[8:]
            price = (msg_split[5])[18:]
            tim = (msg_split[index])[23:-3]
            tim = tim.replace(":", "")
            exit_tim = self.time_to_secs(tim)
            price = float(price)

            for symbl in self.entries.keys():
                if symbl == symbol:
                    difference = price - self.entries[symbol]["Entry Price"]
                    self.entries[symbol]["Exit Price"] = price
                    roi = ((self.entries[symbol]["Exit Price"] - self.entries[symbol]["Entry Price"]) /
                           self.entries[symbol]["Entry Price"]) * 100
                    potential_roi = ((float(highest_price_since_entry) - self.entries[symbol]["Entry Price"]) /
                                     self.entries[symbol]["Entry Price"]) * 100
                    self.entries[symbol]["ROI"] = round(roi,2)
                    self.entries[symbol]["Scan Time"] = scan_time
                    self.entries[symbol]["Exit Time"] = exit_tim
                    self.entries[symbol]["Reason for Exit"] = reason
                    self.entries[symbol]["Human Exit Time"] = human_exit_time
                    self.entries[symbol]["Human Scan Time"] = human_scan_time
                    self.entries[symbol]["Price Diff"] = difference
                    self.entries[symbol]["High Since Entry"] = highest_price_since_entry
                    self.entries[symbol]["Potential ROI"] = potential_roi

        elif "EndOfDayCloseAll" in msg:
            msg_split = msg.split(",")
            index = len(msg_split) - 1

            for i, msg in enumerate(msg_split):
                if "scan_info" in msg:
                    msg_split2 = msg.split("-")
                    scan_tim = msg_split2[2][3:19]
                    human_scan_time = msg_split2[2][3:11]
                    timy = scan_tim.replace(":", "")
                    scan_time = self.time_to_secs(timy)
                elif "human_time" in msg:
                    human_exit_time_eod = msg[23:31]
                elif "high_since_entry" in msg:
                    exclude_last = len(msg) - 2
                    highest_price_since_entry = msg[20:exclude_last]


            symbol = (msg_split[3])[8:]
            price = (msg_split[2])[18:]
            exit_tim = 22800.0
            price = float(price)
            symbol = str(symbol)

            for symbl in self.entries.keys():
                if symbl == symbol:
                    difference = price - self.entries[symbol]["Entry Price"]
                    self.entries[symbol]["Exit Price"] = price
                    roi = ((self.entries[symbol]["Exit Price"] - self.entries[symbol]["Entry Price"]) /
                           self.entries[symbol]["Entry Price"]) * 100
                    potential_roi = ((float(highest_price_since_entry) - self.entries[symbol]["Entry Price"]) /
                           self.entries[symbol]["Entry Price"]) * 100
                    self.entries[symbol]["ROI"] = round(roi, 2)
                    self.entries[symbol]["Scan Time"] = scan_time
                    self.entries[symbol]["Exit Time"] = exit_tim
                    self.entries[symbol]["Reason for Exit"] = "End of Day Close All"
                    self.entries[symbol]["Human Exit Time"] = human_exit_time_eod
                    self.entries[symbol]["Human Scan Time"] = human_scan_time
                    self.entries[symbol]["Price Diff"] = difference
                    self.entries[symbol]["High Since Entry"] = highest_price_since_entry
                    self.entries[symbol]["Potential ROI"] = potential_roi

    def process_msg_date(self, msg):
        if "MARKET TIME" in msg:
            msg_split = msg.split(",")
            index = len(msg_split) - 1

            dia = (msg_split[index])[11:22]
            self.date_ran = dia

    def IB_initial_starting_price(self, msg):
        msg_split = msg.split(",")
        for i in msg_split:
            self.starting_dollar_amount = msg_split[3]
            return

    def IB_entry(self, msg):
        msg_split = msg.split(",")
        symbol = msg_split[1]
        seconds_from_open = int(msg_split[2]) + 34200
        price = float(msg_split[3])
        shares_bought = int(msg_split[4])
        total_money_spent = float(msg_split[5])
        account_balance = float(msg_split[6])
        human_entry_time = str(datetime.timedelta(seconds=seconds_from_open))

        if symbol in self.ib_entries:
            # ToDO: Handel Multiple Entries
            print "DEBUG: Multiple Symbols Found: ", symbol
            time.sleep(5)
        else:
            self.ib_entries[symbol] = {}
            self.ib_entries[symbol].update({'Entry Time': human_entry_time,
                                         "Entry Price": price,
                                         "Shares Purchased": shares_bought,
                                         "Total Spent": total_money_spent,
                                         "Account Balance": account_balance})

    def IB_exit(self, msg):
        msg_split = msg.split(",")
        symbol = msg_split[1]
        seconds_from_open = int(msg_split[2]) + 34200
        price = float(msg_split[3])
        shares_sold = int(msg_split[4])
        total_money_spent = float(msg_split[5])
        account_balance = float(msg_split[6])
        human_exit_time = str(datetime.timedelta(seconds=seconds_from_open))


        if symbol in self.ib_entries:
            difference = total_money_spent - self.ib_entries[symbol]["Total Spent"]
            roi = ((total_money_spent - self.ib_entries[symbol]["Total Spent"]) /
                   self.ib_entries[symbol]["Total Spent"]) * 100
            self.ib_entries[symbol].update({'Exit Time': human_exit_time,
                                         "Exit Price": price,
                                         "Shares Sold": shares_sold,
                                         "Total Spent at Exit": total_money_spent,
                                         "Price Diff": difference,
                                         "ROI": round(roi, 2),
                                         "Account Balance at Exit": account_balance})
        else:
            print "DEBUG: No matching symbol in IB_Entries: ", symbol
            time.sleep(5)

    def dispatcher(self, msg):
        if "EventSignalEnterPosition" in msg:
            self.process_msg_entry(msg)
        elif "EventSignalExitPosition" in msg:
            self.process_msg_exit(msg)
        elif "EndOfDayCloseAll" in msg:
            self.process_msg_exit(msg)
        elif "MARKET TIME" in msg and self.date_ran is None:
            self.process_msg_date(msg)
        elif "|InitialPositionIB|" in msg and self.starting_dollar_amount is None:
            self.IB_initial_starting_price(msg)
        elif "|EnterPositionIB|" in msg:
            self.IB_entry(msg)
        elif "|ExitPositionIB|" in msg:
            self.IB_exit(msg)

    def get_all(self):
        with open(self.path, 'r') as f:
            lines = f.readlines()
            for line in lines:
                if line != "\n" and line != "\r":
                    self.dispatcher(line)


    def create_excel(self):
        # Create a workbook and add a worksheet.
        workbook = xlsxwriter.Workbook('Source_Break_Down_' + str(self.date_ran) + '.xlsx')
        worksheet = workbook.add_worksheet()

        # Add a bold format to use to highlight cells.
        bold = workbook.add_format({'bold': True})
        bold_right = workbook.add_format({'bold': True, 'align': 'right'})
        center = workbook.add_format({'align': 'center'})

        # Add a number format for cells with money.
        money = workbook.add_format({'num_format': '$0.00', 'align': 'left'})
        money_right = workbook.add_format({'num_format': '$0.00', 'align': 'right'})
        percent = workbook.add_format({'align': 'left', 'font_color': 'green'})
        percentred = workbook.add_format({'align': 'left', 'font_color': 'red'})

        # Write some data headers.
        # worksheet.write('A1', 'Item', bold)
        # worksheet.write('B1', 'Cost', bold)

        # Some data we want to write to the worksheet.

        # Start from the first cell below the headers.
        row = 4
        col = 0
        start_titles_on_row = 1

        # Establish Titles
        worksheet.write('A1', 'Date: ' + str(self.date_ran) + " SOURCE BREAK DOWN", bold)

        # Iterate over the data and write it out row by row.
        for source, values in self.accelerant_sources.items():
            worksheet.write(row-1, col, source, bold)
            print source

            worksheet.write(row, 2, 'Symbol', bold)
            worksheet.write(row, 1, 'Act. Bal. After Entry', bold)
            # Entry
            worksheet.write(row, 3, 'Entry Time', bold)
            worksheet.write(row, 4, 'Shares Bought', bold)
            worksheet.write(row, 5, 'Entry Cost ', bold)
            worksheet.write(row, 6, 'Share Price', bold)
            # Exit
            worksheet.write(row, 7, 'Share Price', bold)
            worksheet.write(row, 8, 'Exit Return', bold)
            worksheet.write(row, 9, 'Shares Sold', bold)
            worksheet.write(row, 10, 'Exit Time', bold)
            # Calculations
            worksheet.write(row, 11, 'Act. Bal. After Exit', bold)
            worksheet.write(row, 12, 'Gain/Loss', bold)
            worksheet.write(row, 13, 'ROI %', bold)
            row += 1
            for symbol in values:
                if symbol in self.ib_entries:
                    print symbol
                    worksheet.write(row, 1, self.ib_entries[symbol]['Account Balance'], money)
                    worksheet.write(row, 2, symbol)
                    worksheet.write(row, 3, self.ib_entries[symbol]['Entry Time'])
                    worksheet.write(row, 4, self.ib_entries[symbol]['Shares Purchased'], center)
                    worksheet.write(row, 5, self.ib_entries[symbol]['Total Spent'], money)
                    worksheet.write(row, 6, self.ib_entries[symbol]['Entry Price'], money)
                    worksheet.write(row, 7, self.ib_entries[symbol]['Exit Price'], money)
                    worksheet.write(row, 8, self.ib_entries[symbol]['Total Spent at Exit'], money)
                    worksheet.write(row, 9, self.ib_entries[symbol]['Shares Sold'], center)
                    worksheet.write(row, 10, self.ib_entries[symbol]['Exit Time'])
                    worksheet.write(row, 11, self.ib_entries[symbol]['Account Balance at Exit'], money)
                    worksheet.write(row, 12, self.ib_entries[symbol]['Price Diff'], money_right)
                    # worksheet.write(row, col_sd2 + 10, round(values['Potential ROI'],2))
                    if self.ib_entries[symbol]["ROI"] > 0.0:
                        worksheet.write(row, 13, self.ib_entries[symbol]['ROI'], percent)
                    else:
                        worksheet.write(row, 13, self.ib_entries[symbol]['ROI'], percentred)
                    row += 1

            row += 3

            print "\n"

        workbook.close()




    def execution(self):
        self. get_data_from_accelerant_tracker()
        self.get_all()
        self.create_excel()

        for key, values in self.accelerant_sources.items():
            return
            print key, ":", values, "\n"

        #for key, values in self.entries.items():
            #print key, ":", values, "\n"

        for key, values in self.ib_entries.items():
            return
            print key, ":", values, "\n"



begin = Mixology()
begin.execution()